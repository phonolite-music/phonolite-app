import streamlit as st
import pandas as pd
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.title("Processamento de Royalties")

periodo = st.text_input("Período (AAAA-MM)", placeholder="2024-03")

uploaded_file = st.file_uploader("Carregar arquivo de operações (.xlsx)", type=["xlsx"])
incomes_file = st.file_uploader("Carregar lista de incomes (.xlsx)", type=["xlsx"])

def strip_prefix(name):
    """Remove prefixo do tipo '2026FEV SYNC - ' ou '2026FEV LICENCIAMENTO - ' para normalizar o nome."""
    return re.sub(r'^\d{4}\w+\s+(SYNC|LICENCIAMENTO)\s+-\s+', '', str(name))

def classify(row):
    rh = str(row["Rights-Holder"])
    t = str(row["Type"])
    if t == "Advance Refund":
        return "Recuperação de Adiantamentos"
    if re.search(r"License", t, re.IGNORECASE):
        return "Direitos Autorais – Licenciamento" if rh == "PHONOLITE" else "Repasses Editora - Provisão"
    if re.search(r"Synchro", t, re.IGNORECASE):
        return "Direitos Autorais – Sincronização" if rh == "PHONOLITE" else "Repasses Editora - Provisão"
    if rh == "PHONOLITE":
        return "Direitos Autorais – Fonomecânicos Digitais"
    return "Repasses Editora - Provisão"

if uploaded_file:
    df = pd.read_excel(uploaded_file)

    # Monta payer_map usando nome normalizado (sem prefixo SYNC/LICENCIAMENTO)
    payer_map = {}
    if incomes_file:
        df_incomes = pd.read_excel(incomes_file)
        df_incomes["_base"] = df_incomes["Name"].apply(strip_prefix)
        payer_map = df_incomes.drop_duplicates("_base").set_index("_base")["Payer"].to_dict()

    # --- MÉTRICAS GERAIS ---
    total = df["Amount"].sum()
    st.metric("Total Processado", f"R$ {total:,.2f}")

    st.subheader("Total por Tipo")
    total_por_tipo = df.groupby("Type")["Amount"].sum().reset_index()
    total_por_tipo.columns = ["Tipo", "Total"]
    total_por_tipo["Total"] = total_por_tipo["Total"].map(lambda x: f"R$ {x:,.2f}")
    st.dataframe(total_por_tipo, use_container_width=True, hide_index=True)

    # --- DATAFRAME DE CATEGORIZAÇÃO ---
    st.subheader("Resumo por Categoria Financeira")

    df["Categoria"] = df.apply(classify, axis=1)
    df_resumo = (
        df.groupby(["Categoria", "Name", "Type"])["Amount"]
        .sum()
        .reset_index()
        .rename(columns={"Name": "Nome", "Type": "Tipo", "Amount": "Valor"})
    )

    # Mapeia Fonte via nome normalizado
    df_resumo["_base"] = df_resumo["Nome"].apply(strip_prefix)
    df_resumo["Fonte"] = df_resumo["_base"].map(payer_map).fillna("")
    df_resumo.drop(columns=["_base"], inplace=True)

    if periodo:
        df_resumo.insert(0, "Período", periodo)

    df_display = df_resumo.copy()
    df_display["Valor"] = df_display["Valor"].map(lambda x: f"R$ {x:,.2f}")
    st.dataframe(df_display, use_container_width=True, hide_index=True)

    # --- RESUMO EM TELA ---
    st.subheader("Resumo por Fonte e Categoria")

    df_resumo_tela = df_resumo.copy()
    df_resumo_tela["Fonte"] = df_resumo_tela["Fonte"].fillna("(vazio)").replace("", "(vazio)")
    resumo_agrupado = (
        df_resumo_tela.groupby(["Fonte", "Categoria"])["Valor"]
        .sum()
        .reset_index()
        .sort_values(["Fonte", "Categoria"])
    )

    rows_tela = []
    for fonte in resumo_agrupado["Fonte"].unique():
        grp = resumo_agrupado[resumo_agrupado["Fonte"] == fonte]
        rows_tela.append({"Rótulos de Linha": fonte, "Soma de Valor": grp["Valor"].sum()})
        for _, r in grp.iterrows():
            rows_tela.append({"Rótulos de Linha": f"   {r['Categoria']}", "Soma de Valor": r["Valor"]})
    rows_tela.append({"Rótulos de Linha": "Total Geral", "Soma de Valor": resumo_agrupado["Valor"].sum()})

    df_resumo_display = pd.DataFrame(rows_tela)
    df_resumo_display["Soma de Valor"] = df_resumo_display["Soma de Valor"].map(lambda x: f"R$ {x:,.2f}")
    st.dataframe(df_resumo_display, use_container_width=True, hide_index=True)

    # --- EXPORT XLSX ---
    def write_sheet_plain(ws, df_data):
        """Planilha1: cabeçalho azul escuro/branco bold, linhas pares verde claro, autofilter."""
        df_sorted = df_data.sort_values("Nome") if "Nome" in df_data.columns else df_data

        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill("solid", start_color="1F4E79")
        header_align = Alignment(horizontal="center", vertical="center")
        row_fill_alt = PatternFill("solid", start_color="E2EFDA")
        thin         = Side(style="thin")
        brd          = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = list(df_sorted.columns)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = brd

        for row_idx, row in enumerate(df_sorted.itertuples(index=False), 2):
            fill = row_fill_alt if row_idx % 2 == 0 else PatternFill(fill_type=None)
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.fill   = fill
                cell.border = brd

        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col) + 4
            ws.column_dimensions[col[0].column_letter].width = min(max_len, 50)

    def write_resumo_sheet(ws, df_data):
        """Resumo agrupado: Fonte → Categoria → subtotais → total geral."""
        bold_white = Font(bold=True, color="FFFFFF")
        bold_black = Font(bold=True)
        normal     = Font(bold=False)

        fill_fonte = PatternFill("solid", start_color="1F4E79")
        fill_total = PatternFill("solid", start_color="BDD7EE")
        no_fill    = PatternFill(fill_type=None)

        thin = Side(style="thin")

        def border_row(left=thin, right=thin, top=thin, bottom=thin):
            return Border(left=left, right=right, top=top, bottom=bottom)

        df_r = df_data.copy()
        df_r["Fonte"] = df_r["Fonte"].fillna("(vazio)").replace("", "(vazio)")
        grouped = df_r.groupby(["Fonte", "Categoria"], sort=True)["Valor"].sum()

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 18

        for c, h in enumerate(["Rótulos de Linha", "Soma de Valor"], 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = bold_white
            cell.fill = fill_fonte
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_row(left=Side(style="medium"), right=Side(style="medium"),
                                     top=Side(style="medium"), bottom=Side(style="medium"))

        row_idx     = 2
        total_geral = 0.0

        for fonte, grp in grouped.groupby(level=0):
            subtotal     = grp.sum()
            total_geral += subtotal

            c1 = ws.cell(row=row_idx, column=1, value=fonte)
            c1.font = bold_white; c1.fill = fill_fonte
            c1.border = border_row(left=Side(style="medium"), right=thin,
                                   top=Side(style="medium"), bottom=thin)
            c2 = ws.cell(row=row_idx, column=2, value=subtotal)
            c2.font = bold_white; c2.fill = fill_fonte
            c2.number_format = '#,##0.00'
            c2.border = border_row(left=thin, right=Side(style="medium"),
                                   top=Side(style="medium"), bottom=thin)
            row_idx += 1

            for (_, cat), val in grp.items():
                c1 = ws.cell(row=row_idx, column=1, value=f"   {cat}")
                c1.font = normal; c1.fill = no_fill
                c1.border = border_row(left=Side(style="medium"), right=thin)
                c2 = ws.cell(row=row_idx, column=2, value=val)
                c2.font = normal; c2.fill = no_fill
                c2.number_format = '#,##0.00'
                c2.border = border_row(left=thin, right=Side(style="medium"))
                row_idx += 1

        c1 = ws.cell(row=row_idx, column=1, value="Total Geral")
        c1.font = bold_black; c1.fill = fill_total
        c1.border = border_row(left=Side(style="medium"), right=thin,
                               top=Side(style="medium"), bottom=Side(style="medium"))
        c2 = ws.cell(row=row_idx, column=2, value=total_geral)
        c2.font = bold_black; c2.fill = fill_total
        c2.number_format = '#,##0.00'
        c2.border = border_row(left=thin, right=Side(style="medium"),
                               top=Side(style="medium"), bottom=Side(style="medium"))

    wb  = Workbook()
    ws1 = wb.active
    ws1.title = "Planilha1"
    write_sheet_plain(ws1, df_resumo)

    ws2 = wb.create_sheet("Resumo")
    write_resumo_sheet(ws2, df_resumo)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    fname = f"resumo_royalties_{periodo}.xlsx" if periodo else "resumo_royalties.xlsx"
    st.download_button(
        label="📥 Baixar Resumo para o Financeiro (.xlsx)",
        data=buffer,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
